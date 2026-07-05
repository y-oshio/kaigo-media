#!/usr/bin/env python3
"""令和7年賃金構造基本統計調査 都道府県別第3表(原本XLSX)から対象行を抽出する。

data/kyuryo/raw/chingin-kouzou-r7/pref3-*.xlsx(4分冊)を読み、
男女計ブロックの「介護職員(医療・福祉施設等)」「訪問介護従事者」行を
tidy 形式の CSV(extract.csv)に書き出す。値は原本の表記のまま
(千円・十人の単位換算は build-kyuryo-c2.mjs 側で行う)。

- 出典に値がないセル(「-」等の非数値)は空欄で出力する(null 化は build 側)。
- 手入力・推計は一切行わない(決定的変換のみ)。
- 依存: openpyxl(pip install openpyxl)

使い方: python3 scripts/extract-kyuryo-chingin-r7.py
出力:   data/kyuryo/raw/chingin-kouzou-r7/extract.csv
"""

import csv
import re
import unicodedata
from pathlib import Path

import openpyxl

RAW_DIR = Path(__file__).resolve().parent.parent / "data" / "kyuryo" / "raw" / "chingin-kouzou-r7"
OUT_CSV = RAW_DIR / "extract.csv"

# 統計表上の職種名(正規化後) → 本サイトの jobSlug(data/kb/jobs.ts と同期)
TARGET_JOBS = {
    "介護職員(医療・福祉施設等)": "kaigoshoku",
    "訪問介護従事者": "homonkaigo",
}

# 表頭の8列(全国・各都道府県ブロック共通の並び。原本の列順を変えない)
COLUMNS = [
    "age",              # 年齢(歳)
    "tenureYears",      # 勤続年数(年)
    "scheduledHours",   # 所定内実労働時間数(時間)
    "overtimeHours",    # 超過実労働時間数(時間)
    "monthlyKimatte",   # きまって支給する現金給与額(千円)
    "monthlyShotei",    # 所定内給与額(千円)
    "annualBonus",      # 年間賞与その他特別給与額(千円)
    "workersTens",      # 労働者数(十人)
]


def normalize(text: str) -> str:
    """職種名の照合用正規化(全角→半角括弧・改行/空白除去)。"""
    t = unicodedata.normalize("NFKC", str(text))
    return re.sub(r"\s+", "", t)


def pref_header(value: str) -> str | None:
    """表頭セルから都道府県名を取り出す。'０１北海道'→'01,北海道'、'全国'→'00,全国'。"""
    t = normalize(value)
    if t == "全国":
        return "00,全国"
    m = re.fullmatch(r"(\d{2})(\S+)", t)
    return f"{m.group(1)},{m.group(2)}" if m else None


def cell_number(v):
    """数値セルのみ値を返す。'-' 等の非数値(出典に値がない)は None。"""
    return v if isinstance(v, (int, float)) else None


rows_out = []
for xlsx in sorted(RAW_DIR.glob("pref3-*.xlsx")):
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    ws = wb[wb.sheetnames[0]]
    data = [list(r) for r in ws.iter_rows(values_only=True)]

    # 表頭(6行目)から都道府県ブロックの開始列を得る
    header = data[5]
    blocks = []  # (開始列index, '県コード,県名')
    for idx, v in enumerate(header):
        if idx >= 3 and v not in (None, ""):
            p = pref_header(v)
            if p:
                blocks.append((idx, p))
    if not blocks:
        raise SystemExit(f"{xlsx.name}: 表頭に都道府県が見つかりません(表構造が変わった可能性)")

    # 職種行: 最初の出現 = 男女計ブロック(男・女ブロックは2回目以降)
    found = {}
    for row in data:
        label = normalize(row[1]) if row[1] else ""
        for jp_name, slug in TARGET_JOBS.items():
            # 先頭行は「男女計\n\n職種名」のように性別見出しが同居するため末尾一致で照合
            if slug not in found and label.endswith(normalize(jp_name)):
                found[slug] = row
    missing = set(TARGET_JOBS.values()) - set(found)
    if missing:
        raise SystemExit(f"{xlsx.name}: 対象職種行が見つかりません: {missing}")

    for slug, row in found.items():
        for start, pref in blocks.items() if isinstance(blocks, dict) else blocks:
            values = [cell_number(row[start + i]) for i in range(len(COLUMNS))]
            code, name = pref.split(",")
            rows_out.append([xlsx.name, slug, code, name] + values)

# 全国(00)→県コード順、職種順で安定ソート
rows_out.sort(key=lambda r: (r[2], r[1]))

with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["file", "jobSlug", "prefCode", "prefName"] + COLUMNS)
    for r in rows_out:
        w.writerow(["" if v is None else v for v in r])

prefs = {r[2] for r in rows_out}
print(f"extract: {OUT_CSV.relative_to(RAW_DIR.parent.parent.parent.parent)}")
print(f"  行数 {len(rows_out)}(職種2 × 地域 {len(prefs)}: 全国+{len(prefs) - 1}都道府県)")
if len(prefs) != 48:
    raise SystemExit("NG: 全国+47都道府県になっていません")
